#!/usr/bin/env python3
"""
GitHub Starred Repositories Tracker
Fetches all starred repositories and generates README.md, Excel, and Interactive HTML automatically!
"""

import os
import sys
import json
import time
import requests
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

GITHUB_TOKEN = os.getenv('GITHUB_TOKEN')
if not GITHUB_TOKEN:
    print("Error: GITHUB_TOKEN not found in .env file")
    sys.exit(1)

HEADERS = {
    'Authorization': f'token {GITHUB_TOKEN}',
    'Accept': 'application/vnd.github.v3+json'
}

# Constants for retry logic
MAX_RETRIES = 5
INITIAL_BACKOFF = 2  # seconds

def make_request_with_retry(url: str, headers: Dict[str, str]) -> Optional[requests.Response]:
    """Make a request with retry logic and rate limit handling"""
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.get(url, headers=headers, timeout=30)

            # Check rate limit
            remaining = int(response.headers.get('X-RateLimit-Remaining', 1))
            if remaining == 0:
                reset_time = int(response.headers.get('X-RateLimit-Reset', 0))
                wait_time = max(reset_time - int(time.time()), 0) + 1
                print(f"  Rate limit hit. Waiting {wait_time} seconds...")
                time.sleep(wait_time)
                continue

            # Handle rate limit response
            if response.status_code == 403 and 'rate limit' in response.text.lower():
                reset_time = int(response.headers.get('X-RateLimit-Reset', 0))
                wait_time = max(reset_time - int(time.time()), 60) + 1
                print(f"  Rate limit exceeded. Waiting {wait_time} seconds...")
                time.sleep(wait_time)
                continue

            # Success
            if response.status_code == 200:
                return response

            # Server error - retry
            if response.status_code >= 500:
                backoff = INITIAL_BACKOFF * (2 ** attempt)
                print(f"  Server error {response.status_code}. Retrying in {backoff}s (attempt {attempt + 1}/{MAX_RETRIES})")
                time.sleep(backoff)
                continue

            # Client error - don't retry
            print(f"Error: GitHub API returned status code {response.status_code}")
            print(response.text)
            return None

        except requests.exceptions.Timeout:
            backoff = INITIAL_BACKOFF * (2 ** attempt)
            print(f"  Request timeout. Retrying in {backoff}s (attempt {attempt + 1}/{MAX_RETRIES})")
            time.sleep(backoff)
        except requests.exceptions.RequestException as e:
            backoff = INITIAL_BACKOFF * (2 ** attempt)
            print(f"  Request error: {e}. Retrying in {backoff}s (attempt {attempt + 1}/{MAX_RETRIES})")
            time.sleep(backoff)

    print(f"Failed to fetch after {MAX_RETRIES} attempts")
    return None


def parse_link_header(link_header: str) -> Dict[str, str]:
    """Parse the Link header to extract pagination URLs"""
    links = {}
    if not link_header:
        return links

    for part in link_header.split(','):
        section = part.split(';')
        if len(section) < 2:
            continue
        url = section[0].strip()[1:-1]  # Remove < and >
        rel = section[1].strip().split('=')[1].strip('"')
        links[rel] = url

    return links


def fetch_all_starred_repos() -> List[Dict[str, Any]]:
    """Fetch all starred repositories with robust pagination using Link header"""
    print("Fetching starred repositories...")
    all_repos = []
    per_page = 100

    # Use Link header pagination (more reliable than manual page counting)
    url = f'https://api.github.com/user/starred?per_page={per_page}'
    page = 1

    while url:
        response = make_request_with_retry(url, HEADERS)

        if response is None:
            print("Error: Failed to fetch starred repositories")
            sys.exit(1)

        repos = response.json()
        if not repos:
            break

        all_repos.extend(repos)
        print(f"  Fetched page {page} ({len(repos)} repos, total: {len(all_repos)})")

        # Get next page URL from Link header
        link_header = response.headers.get('Link', '')
        links = parse_link_header(link_header)
        url = links.get('next')
        page += 1

        # Small delay to be nice to the API
        time.sleep(0.1)

    print(f"Total repositories fetched: {len(all_repos)}")
    return all_repos

def categorize_repo(repo: Dict[str, Any]) -> str:
    """Categorize repository based on language and topics"""
    language = repo.get('language', '')
    topics = repo.get('topics', [])

    # Define category mappings
    web_langs = ['JavaScript', 'TypeScript', 'HTML', 'CSS', 'Vue', 'React']
    backend_langs = ['Python', 'Java', 'Go', 'Ruby', 'PHP', 'C#', 'Rust', 'Kotlin']
    mobile_langs = ['Swift', 'Objective-C', 'Dart', 'Kotlin', 'Java']
    data_langs = ['Python', 'R', 'Julia', 'Jupyter Notebook']

    # Check topics first for more specific categorization
    topic_str = ' '.join(topics).lower()

    if any(t in topic_str for t in ['machine-learning', 'deep-learning', 'ai', 'ml', 'neural-network', 'data-science']):
        return 'AI/ML'
    elif any(t in topic_str for t in ['devops', 'docker', 'kubernetes', 'cicd', 'ci-cd']):
        return 'DevOps'
    elif any(t in topic_str for t in ['security', 'cybersecurity', 'penetration-testing']):
        return 'Security'
    elif any(t in topic_str for t in ['frontend', 'react', 'vue', 'angular', 'web']):
        return 'Frontend'
    elif any(t in topic_str for t in ['backend', 'api', 'server']):
        return 'Backend'
    elif any(t in topic_str for t in ['mobile', 'android', 'ios', 'flutter']):
        return 'Mobile'
    elif any(t in topic_str for t in ['cli', 'command-line', 'terminal']):
        return 'CLI Tools'
    elif any(t in topic_str for t in ['database', 'sql', 'nosql']):
        return 'Database'
    elif any(t in topic_str for t in ['blockchain', 'cryptocurrency', 'web3']):
        return 'Blockchain'

    # Fallback to language-based categorization
    if language in web_langs:
        return 'Web Development'
    elif language in backend_langs:
        return 'Backend'
    elif language in mobile_langs:
        return 'Mobile'
    elif language == 'Python' and any(t in topic_str for t in ['data', 'analysis', 'science']):
        return 'Data Science'
    elif language:
        return language
    else:
        return 'Other'

def truncate_description(description: str, max_sentences: int = 3) -> str:
    """Truncate description to max sentences"""
    if not description:
        return "No description available"

    # Simple sentence splitting
    sentences = description.replace('! ', '!|').replace('? ', '?|').replace('. ', '.|').split('|')
    truncated = ' '.join(sentences[:max_sentences]).strip()

    # Limit to reasonable length
    if len(truncated) > 200:
        truncated = truncated[:197] + '...'

    return truncated

def format_date(date_str: str) -> str:
    """Format ISO date to DD-MM-YYYY HH:MM"""
    if not date_str:
        return "N/A"

    try:
        dt = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%SZ')
        return dt.strftime('%d-%m-%Y %H:%M')
    except:
        return date_str

def process_repos(repos: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Process repositories into structured data"""
    print("Processing repositories...")
    processed = []

    for repo in repos:
        processed.append({
            'name': repo['full_name'],
            'url': repo['html_url'],
            'category': categorize_repo(repo),
            'description': truncate_description(repo.get('description', '')),
            'stars': repo.get('stargazers_count', 0),
            'language': repo.get('language', 'N/A'),
            'last_updated': format_date(repo.get('updated_at', '')),
            'created_at': format_date(repo.get('created_at', '')),
            'topics': ', '.join(repo.get('topics', [])[:5])  # First 5 topics
        })

    # Sort by stars by default (descending)
    processed.sort(key=lambda x: x['stars'], reverse=True)

    print(f"Processed {len(processed)} repositories")
    return processed

def generate_interactive_html(repos: List[Dict[str, str]]) -> None:
    """Generate index.html for GitHub Pages with sortable table and search"""
    print("Generating index.html for GitHub Pages...")

    last_updated = datetime.utcnow().strftime('%d-%m-%Y %H:%M UTC')
    total_repos = len(repos)
    total_stars = sum(r['stars'] for r in repos)

    categories = {}
    for repo in repos:
        cat = repo['category']
        categories[cat] = categories.get(cat, 0) + 1

    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GitHub Starred Repositories</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: #000000;
            color: #e6e6e6;
            line-height: 1.6;
            overflow-x: hidden;
        }}

        /* Compact Header */
        .header {{
            background: linear-gradient(135deg, #000000 0%, #1a1a1a 100%);
            padding: 1.5rem 2rem;
            border-bottom: 1px solid #222;
            position: sticky;
            top: 0;
            z-index: 100;
            backdrop-filter: blur(10px);
        }}

        .header-content {{
            max-width: 1600px;
            margin: 0 auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 2rem;
            flex-wrap: wrap;
        }}

        .title {{
            display: flex;
            align-items: center;
            gap: 0.75rem;
        }}

        .title h1 {{
            font-size: 1.5rem;
            font-weight: 700;
            background: linear-gradient(135deg, #60a5fa 0%, #a78bfa 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}

        .title .emoji {{
            font-size: 1.75rem;
        }}

        /* Stats Bar - Inline */
        .stats-inline {{
            display: flex;
            gap: 2rem;
            align-items: center;
            font-size: 0.875rem;
        }}

        .stat-inline {{
            display: flex;
            flex-direction: column;
            align-items: center;
        }}

        .stat-number {{
            font-size: 1.25rem;
            font-weight: 700;
            color: #60a5fa;
        }}

        .stat-label {{
            font-size: 0.75rem;
            color: #888;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        /* Search Bar */
        .search-container {{
            max-width: 1600px;
            margin: 1.5rem auto;
            padding: 0 2rem;
            position: relative;
            z-index: 50;
            background: #000000;
        }}

        .search-wrapper {{
            position: relative;
            max-width: 600px;
            margin: 0 auto;
        }}

        .search-icon {{
            position: absolute;
            left: 1rem;
            top: 50%;
            transform: translateY(-50%);
            font-size: 1.125rem;
            opacity: 0.5;
        }}

        #searchInput {{
            width: 100%;
            padding: 0.875rem 1rem 0.875rem 3rem;
            background: #0a0a0a;
            border: 1px solid #222;
            border-radius: 12px;
            color: #e6e6e6;
            font-size: 0.9375rem;
            font-family: 'Inter', sans-serif;
            transition: all 0.3s ease;
        }}

        #searchInput:focus {{
            outline: none;
            border-color: #60a5fa;
            background: #111;
            box-shadow: 0 0 0 3px rgba(96, 165, 250, 0.1);
        }}

        #searchInput::placeholder {{
            color: #555;
        }}

        /* Table Container */
        .table-container {{
            max-width: 1600px;
            margin: 0 auto;
            padding: 0 2rem 2rem;
        }}

        table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            background: #0a0a0a;
            border-radius: 12px;
            overflow: hidden;
            border: 1px solid #1a1a1a;
        }}

        thead {{
            background: #111;
        }}

        th {{
            padding: 1rem 1.25rem;
            text-align: left;
            font-weight: 600;
            font-size: 0.8125rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: #888;
            cursor: pointer;
            user-select: none;
            white-space: nowrap;
            border-bottom: 1px solid #1a1a1a;
            transition: all 0.2s ease;
        }}

        th:hover {{
            background: #1a1a1a;
            color: #60a5fa;
        }}

        .sort-indicator {{
            display: inline-block;
            margin-left: 0.375rem;
            font-size: 0.75rem;
            opacity: 0.6;
            transition: all 0.2s ease;
        }}

        th:hover .sort-indicator {{
            opacity: 1;
        }}

        td {{
            padding: 1.125rem 1.25rem;
            border-bottom: 1px solid #0f0f0f;
            font-size: 0.875rem;
            word-wrap: break-word;
        }}

        tbody tr {{
            transition: all 0.2s ease;
        }}

        tbody tr:hover {{
            background: #0f0f0f;
        }}

        tbody tr:last-child td {{
            border-bottom: none;
        }}

        a {{
            color: #60a5fa;
            text-decoration: none;
            font-weight: 500;
            transition: color 0.2s ease;
        }}

        a:hover {{
            color: #93c5fd;
            text-decoration: underline;
        }}

        /* Column widths */
        th:nth-child(1), td:nth-child(1) {{ width: 22%; }}
        th:nth-child(2), td:nth-child(2) {{ width: 11%; text-align: center; }}
        th:nth-child(3), td:nth-child(3) {{ width: 9%; text-align: center; }}
        th:nth-child(4), td:nth-child(4) {{ width: 10%; text-align: center; }}
        th:nth-child(5), td:nth-child(5) {{ width: 36%; }}
        th:nth-child(6), td:nth-child(6) {{ width: 12%; text-align: center; font-size: 0.8125rem; color: #666; }}

        /* Category badges */
        .category-badge {{
            display: inline-block;
            padding: 0.25rem 0.625rem;
            background: #1a1a1a;
            border-radius: 6px;
            font-size: 0.75rem;
            font-weight: 500;
            border: 1px solid #222;
        }}

        /* Language badges */
        .language-badge {{
            display: inline-block;
            padding: 0.25rem 0.625rem;
            background: linear-gradient(135deg, #1a1a1a 0%, #0f0f0f 100%);
            border-radius: 6px;
            font-size: 0.75rem;
            font-weight: 500;
            border: 1px solid #1a1a1a;
        }}

        /* Stars */
        .stars {{
            font-weight: 600;
            color: #fbbf24;
        }}

        /* Responsive */
        @media (max-width: 768px) {{
            .header {{
                padding: 1rem;
            }}

            .header-content {{
                flex-direction: column;
                align-items: flex-start;
                gap: 1rem;
            }}

            .stats-inline {{
                gap: 1rem;
                font-size: 0.8125rem;
            }}

            .search-container, .table-container {{
                padding: 1rem;
            }}

            th, td {{
                padding: 0.75rem 0.875rem;
                font-size: 0.8125rem;
            }}
        }}

        /* Scrollbar */
        ::-webkit-scrollbar {{
            width: 8px;
            height: 8px;
        }}

        ::-webkit-scrollbar-track {{
            background: #000;
        }}

        ::-webkit-scrollbar-thumb {{
            background: #222;
            border-radius: 4px;
        }}

        ::-webkit-scrollbar-thumb:hover {{
            background: #333;
        }}

        /* Back to Top Button */
        .back-to-top {{
            position: fixed;
            bottom: 1.5rem;
            right: 1.5rem;
            width: 38px;
            height: 38px;
            background: rgba(15, 15, 15, 0.9);
            border: 1px solid rgba(96, 165, 250, 0.4);
            border-radius: 50%;
            color: #60a5fa;
            cursor: pointer;
            opacity: 1;
            visibility: visible;
            transition: all 0.25s ease;
            z-index: 99999;
            display: flex;
            align-items: center;
            justify-content: center;
            backdrop-filter: blur(8px);
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.4);
        }}

        .back-to-top:hover {{
            background: rgba(96, 165, 250, 0.15);
            border-color: #60a5fa;
            transform: scale(1.1);
            box-shadow: 0 4px 16px rgba(96, 165, 250, 0.3);
        }}

        .back-to-top:active {{
            transform: scale(0.95);
        }}

        @media (max-width: 768px) {{
            .back-to-top {{
                bottom: 1rem;
                right: 1rem;
                width: 34px;
                height: 34px;
            }}
            .back-to-top svg {{
                width: 16px;
                height: 16px;
            }}
        }}
    </style>
</head>
<body>
    <header class="header">
        <div class="header-content">
            <div class="title">
                <span class="emoji">⭐</span>
                <h1>GitHub Starred Repositories</h1>
            </div>
            <div class="stats-inline">
                <div class="stat-inline">
                    <div class="stat-number">{total_repos:,}</div>
                    <div class="stat-label">Repos</div>
                </div>
                <div class="stat-inline">
                    <div class="stat-number">{total_stars:,}</div>
                    <div class="stat-label">Stars</div>
                </div>
                <div class="stat-inline">
                    <div class="stat-number">{len(categories)}</div>
                    <div class="stat-label">Categories</div>
                </div>
                <div class="stat-inline">
                    <div class="stat-number">{last_updated}</div>
                    <div class="stat-label">Updated</div>
                </div>
            </div>
        </div>
    </header>

    <div class="search-container">
        <div class="search-wrapper">
            <span class="search-icon">🔍</span>
            <input type="text" id="searchInput" placeholder="Search repositories..." onkeyup="searchTable()">
        </div>
    </div>

    <div class="table-container">
        <table id="repoTable">
            <thead>
                <tr>
                    <th onclick="sortTable(0)">Repository <span id="sort0" class="sort-indicator">🔽</span></th>
                    <th onclick="sortTable(1)">Category <span id="sort1" class="sort-indicator">🔽</span></th>
                    <th onclick="sortTable(2)">Stars <span id="sort2" class="sort-indicator">🔽</span></th>
                    <th onclick="sortTable(3)">Language <span id="sort3" class="sort-indicator">🔽</span></th>
                    <th>Description</th>
                    <th onclick="sortTable(5)">Updated <span id="sort5" class="sort-indicator">🔽</span></th>
                </tr>
            </thead>
            <tbody>
'''

    # Add repository rows with styled badges
    for repo in repos:
        html_content += f'''            <tr>
                <td><a href="{repo['url']}" target="_blank">{repo['name']}</a></td>
                <td><span class="category-badge">{repo['category']}</span></td>
                <td><span class="stars">{repo['stars']:,}</span></td>
                <td><span class="language-badge">{repo['language']}</span></td>
                <td>{repo['description']}</td>
                <td>{repo['last_updated']}</td>
            </tr>
'''

    html_content += '''        </tbody>
    </table>

    <script>
        let sortDirection = {};

        function sortTable(columnIndex) {
            const table = document.getElementById("repoTable");
            const tbody = table.querySelector("tbody");
            const rows = Array.from(tbody.querySelectorAll("tr"));

            // Toggle sort direction
            sortDirection[columnIndex] = !sortDirection[columnIndex];
            const ascending = sortDirection[columnIndex];

            rows.sort((a, b) => {
                let aValue = a.cells[columnIndex].textContent.trim();
                let bValue = b.cells[columnIndex].textContent.trim();

                // Handle numeric sorting for stars
                if (columnIndex === 2) {
                    aValue = parseInt(aValue.replace(/,/g, ''));
                    bValue = parseInt(bValue.replace(/,/g, ''));
                }

                // Handle date sorting
                if (columnIndex === 5) {
                    const parseDate = (dateStr) => {
                        if (dateStr === 'N/A') return new Date(0);
                        const [date, time] = dateStr.split(' ');
                        if (!date) return new Date(0);
                        const [day, month, year] = date.split('-');
                        return new Date(`${year}-${month}-${day}T${time || '00:00'}:00`);
                    };
                    aValue = parseDate(aValue);
                    bValue = parseDate(bValue);
                }

                if (aValue < bValue) return ascending ? -1 : 1;
                if (aValue > bValue) return ascending ? 1 : -1;
                return 0;
            });

            // Re-append sorted rows
            rows.forEach(row => tbody.appendChild(row));

            // Update sort indicators
            document.querySelectorAll('.sort-indicator').forEach(span => {
                span.textContent = '🔽';
            });
            document.getElementById(`sort${columnIndex}`).textContent = ascending ? '🔼' : '🔽';
        }

        function searchTable() {
            const input = document.getElementById("searchInput");
            const filter = input.value.toLowerCase();
            const table = document.getElementById("repoTable");
            const tr = table.getElementsByTagName("tr");

            for (let i = 1; i < tr.length; i++) {
                const row = tr[i];
                const text = row.textContent.toLowerCase();
                row.style.display = text.includes(filter) ? "" : "none";
            }
        }

        // Initialize - sort by stars (already sorted in Python, but show indicator)
        sortDirection[2] = false; // Will toggle to true on first click
    </script>

    <!-- Back to Top Button -->
    <button id="backToTop" class="back-to-top" title="Back to top" onclick="window.scrollTo({{top: 0, behavior: 'smooth'}})">
        <svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
            <path d="M18 15l-6-6-6 6"/>
        </svg>
    </button>
</body>
</html>
'''

    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html_content)

    print("index.html generated successfully for GitHub Pages!")

def generate_readme(repos: List[Dict[str, str]]) -> None:
    """Generate README.md with properly formatted table"""
    print("Generating README.md...")

    # Calculate statistics
    total_repos = len(repos)
    total_stars = sum(r['stars'] for r in repos)
    categories = {}
    for repo in repos:
        cat = repo['category']
        categories[cat] = categories.get(cat, 0) + 1

    # Get last updated time
    last_updated = datetime.utcnow().strftime('%d-%m-%Y %H:%M UTC')

    readme_content = f'''# 🌟 My GitHub Starred Repositories

> List of all my starred repositories on GitHub, updated and generated automatically!

## 📊 Statistics

- **Total Repositories**: {total_repos:,}
- **Total Stars Given**: {total_stars:,}
- **Categories**: {len(categories)}
- **Last Updated**: {last_updated}

### 📁 Category Distribution

'''

    # Add category statistics
    for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_repos) * 100
        readme_content += f"- **{cat}**: {count} repos ({percentage:.1f}%)\n"

    # Show top 50 repos in README to avoid cluttering
    display_limit = 50

    readme_content += f'''
---

## 📥 Interactive Versions

For the full experience with **sortable columns** and **live search functionality**:

- 🌐 **[Live GitHub Pages](https://github.ragilmalik.com/starred-repos)** - Interactive web app with sorting and real-time search
- 📊 **[Excel Spreadsheet](./starred_repos.xlsx)** - Full data with filters and sorting
- 💾 **[Download HTML](./index.html)** - Offline version you can open locally

> **Note**: GitHub doesn't support JavaScript in README files, so the table below is static.
> Visit the GitHub Pages link above for the full interactive experience!

---

## 🗂️ Top {display_limit} Starred Repositories (by stars)

<table>
<thead>
<tr>
<th align="left">Repository</th>
<th align="center">Category</th>
<th align="center">Stars ⭐</th>
<th align="center">Language</th>
<th align="left">Description</th>
<th align="center">Last Updated</th>
</tr>
</thead>
<tbody>
'''

    # Add repository rows (limited to display_limit)
    for repo in repos[:display_limit]:
        # Truncate description for README display
        desc = repo['description']
        if len(desc) > 100:
            desc = desc[:97] + '...'

        readme_content += f'''<tr>
<td><a href="{repo['url']}">{repo['name']}</a></td>
<td align="center">{repo['category']}</td>
<td align="center">{repo['stars']:,}</td>
<td align="center">{repo['language']}</td>
<td>{desc}</td>
<td align="center"><sub>{repo['last_updated']}</sub></td>
</tr>
'''

    readme_content += f'''</tbody>
</table>

<details>
<summary><b>📋 View All {total_repos} Repositories by Category</b></summary>
<br>

'''

    # Group repositories by category
    by_category = {}
    for repo in repos:
        cat = repo['category']
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(repo)

    # Display repositories grouped by category
    for cat in sorted(by_category.keys()):
        cat_repos = by_category[cat]
        readme_content += f'''### {cat} ({len(cat_repos)} repositories)

'''
        for repo in cat_repos[:20]:  # Show max 20 per category
            desc_short = repo['description'][:100] + ("..." if len(repo['description']) > 100 else "")
            readme_content += f'''- [{repo['name']}]({repo['url']}) ⭐ {repo['stars']:,} - {desc_short}\n'''

        if len(cat_repos) > 20:
            readme_content += f'''\n*...and {len(cat_repos) - 20} more. See [GitHub Pages](https://ragilmalik.github.io/stars-repos/) or [Excel file](./starred_repos.xlsx) for complete list.*\n'''

        readme_content += '\n'

    readme_content += f'''
</details>

---

## 🔄 Automatic Updates

This repository is automatically updated **daily at 00:00 UTC** using GitHub Actions.

**What gets updated:**
- ✅ All {total_repos:,} starred repositories (with pagination support)
- ✅ Star counts and last updated time
- ✅ All newly added repositories
- ✅ Category statistics and analytics
- ✅ README.md (generated automatically), Excel file, and Interactive HTML

---

## 👥 Want to use this for your own starred repos?

Check out the **[TUTORIAL.md](./TUTORIAL.md)** for complete setup instructions!

---

<div align="center">

**Made with ❤️ using Python and GitHub Actions**

*Last generated: {last_updated}*

[🌐 Live Interactive View](https://github.ragilmalik.com/starred-repos) • [📥 Download Excel](./starred_repos.xlsx) • [📖 Setup Tutorial](./TUTORIAL.md)

</div>
'''

    with open('README.md', 'w', encoding='utf-8') as f:
        f.write(readme_content)

    print("README.md generated successfully!")

def generate_excel(repos: List[Dict[str, str]]) -> None:
    """Generate Excel file with repository data"""
    print("Generating Excel file...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Starred Repositories"

    # Define headers
    headers = ['Repository Name', 'URL', 'Category', 'Description', 'Stars', 'Language', 'Last Updated', 'Created At', 'Topics']

    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data
    for row_idx, repo in enumerate(repos, 2):
        ws.cell(row=row_idx, column=1, value=repo['name'])
        ws.cell(row=row_idx, column=2, value=repo['url'])
        ws.cell(row=row_idx, column=3, value=repo['category'])
        ws.cell(row=row_idx, column=4, value=repo['description'])
        ws.cell(row=row_idx, column=5, value=repo['stars'])
        ws.cell(row=row_idx, column=6, value=repo['language'])
        ws.cell(row=row_idx, column=7, value=repo['last_updated'])
        ws.cell(row=row_idx, column=8, value=repo['created_at'])
        ws.cell(row=row_idx, column=9, value=repo['topics'])

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Enable filters
    ws.auto_filter.ref = ws.dimensions

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save('starred_repos.xlsx')
    print("Excel file generated successfully!")

def main():
    """Main execution function"""
    print("=" * 60)
    print("GitHub Starred Repositories Tracker")
    print("=" * 60)

    # Fetch repositories
    repos = fetch_all_starred_repos()

    # Process data
    processed_repos = process_repos(repos)

    # Generate outputs
    generate_readme(processed_repos)
    generate_interactive_html(processed_repos)
    generate_excel(processed_repos)

    print("=" * 60)
    print("✅ All done! Files generated:")
    print("  - README.md (overview)")
    print("  - index.html (GitHub Pages - interactive)")
    print("  - starred_repos.xlsx (Excel data)")
    print("=" * 60)

if __name__ == '__main__':
    main()
